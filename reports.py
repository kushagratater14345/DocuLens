"""
reports.py - DocuLens AI report export

Generates downloadable Excel and PDF summaries of everything currently
stored in ChromaDB: document list + analytics.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import core

REPORTS_FOLDER = core.DATA_FOLDER / "reports"
REPORTS_FOLDER.mkdir(parents=True, exist_ok=True)


def _document_rows(store: "core.DocumentStore") -> List[Dict[str, Any]]:
    docs = store.get_all()
    rows = []
    for doc in docs:
        meta = doc["metadata"]
        rows.append(
            {
                "Filename": meta.get("source", ""),
                "Type": meta.get("document_type", ""),
                "Category": meta.get("category", ""),
                "Vendor": meta.get("vendor") or meta.get("bank_name") or meta.get("organization") or "",
                "Invoice Number": meta.get("invoice_number", ""),
                "Date": meta.get("date", ""),
                "Total Amount": core._to_float(meta.get("total_amount")) or "",
                "GST": core._to_float(meta.get("gst")) or "",
                "Currency": meta.get("currency", ""),
                "Confidence": meta.get("confidence", ""),
                "Risk Level": meta.get("risk_level", ""),
                "Processed At": meta.get("processed_at", ""),
            }
        )
    return rows


def generate_excel_report(store: "core.DocumentStore") -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = _document_rows(store)
    stats = store.compute_stats()

    wb = Workbook()

    # --- Documents sheet ---
    ws = wb.active
    ws.title = "Documents"

    headers = list(rows[0].keys()) if rows else [
        "Filename", "Type", "Category", "Vendor", "Invoice Number", "Date",
        "Total Amount", "GST", "Currency", "Confidence", "Risk Level", "Processed At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)

    # --- Summary sheet ---
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    summary_rows = [
        ("Total Documents", stats.get("documents", 0)),
        ("Total Spending", stats.get("total_spending", 0)),
        ("Total GST", stats.get("total_gst", 0)),
        ("High Risk Documents", stats.get("high_risk", 0)),
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in summary_rows:
        summary.append([label, value])

    summary.append([])
    summary.append(["By Document Type", ""])
    for k, v in stats.get("by_type", {}).items():
        summary.append([k, v])

    summary.append([])
    summary.append(["By Category (Spending)", ""])
    for k, v in stats.get("by_category", {}).items():
        summary.append([k, v])

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 20

    out_path = REPORTS_FOLDER / "doculens_report.xlsx"
    wb.save(out_path)
    return out_path


def generate_pdf_report(store: "core.DocumentStore") -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    rows = _document_rows(store)
    stats = store.compute_stats()
    styles = getSampleStyleSheet()

    out_path = REPORTS_FOLDER / "doculens_report.pdf"
    doc = SimpleDocTemplate(str(out_path), pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    story = []
    story.append(Paragraph("DocuLens AI — Document Analytics Report", styles["Title"]))
    story.append(Paragraph(datetime.now().strftime("Generated on %B %d, %Y at %H:%M"), styles["Normal"]))
    story.append(Spacer(1, 16))

    story.append(Paragraph("Summary", styles["Heading2"]))
    summary_data = [
        ["Total Documents", str(stats.get("documents", 0))],
        ["Total Spending", f"{stats.get('total_spending', 0):,}"],
        ["Total GST", f"{stats.get('total_gst', 0):,}"],
        ["High Risk Documents", str(stats.get("high_risk", 0))],
    ]
    summary_table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF0FB")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Documents", styles["Heading2"]))
    if rows:
        table_data = [["Vendor", "Type", "Date", "Total", "GST", "Risk"]]
        for r in rows:
            table_data.append(
                [
                    str(r.get("Vendor", ""))[:24],
                    str(r.get("Type", "")),
                    str(r.get("Date", "")),
                    str(r.get("Total Amount", "")),
                    str(r.get("GST", "")),
                    str(r.get("Risk Level", "")),
                ]
            )
        doc_table = Table(table_data, colWidths=[4 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2 * cm])
        doc_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(doc_table)
    else:
        story.append(Paragraph("No documents processed yet.", styles["Normal"]))

    doc.build(story)
    return out_path
